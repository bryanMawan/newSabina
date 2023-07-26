from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def add_row_to_excel(file_path, data):
    # Load the workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook(file_path)
        print("Workbook loaded.")
    except FileNotFoundError:
        workbook = Workbook()
        print("New workbook created.")

    # Select the active sheet
    sheet = workbook.active

    # Create headers if the sheet is empty
    if sheet.max_row == 0:
        headers = list(data.keys())
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = data[header]
            sheet[f"{col_letter}1"].font = Font(bold=True)
        print("Headers added.")

    # Append the data to a new row
    row_data = list(data.values())
    new_row = sheet.max_row + 1
    for col_num, value in enumerate(row_data, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}{new_row}"] = value
    print("Data added.")

    # Save the workbook
    workbook.save(file_path)
    workbook.close()
    print("Workbook saved.")


# Example usage
data = {
    "name": "John Doe",
    "age": 30,
    "stadsdel": "Stadsområde Centrum",
    "idrott": "Football",
    "önskad idrott": "Tennis",
    "in a union": True
}
excel_password = "Soptunna1!"
